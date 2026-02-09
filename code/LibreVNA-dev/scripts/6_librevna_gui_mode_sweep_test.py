#!/usr/bin/env python3
"""
6_librevna_gui_mode_sweep_test.py
-----------------------------------
Unified single-sweep / continuous-sweep benchmark for LibreVNA.

Reads sweep parameters from sweep_config.yaml, runs the requested mode
(single or continuous) across every IFBW value listed in the config, and
produces both a console PrettyTable summary and a multi-sheet xlsx workbook.

Class hierarchy
---------------
    BaseVNASweep            (ABC)  -- shared lifecycle, GUI, cal, xlsx
    SingleModeSweep         (BaseVNASweep)  -- single-sweep trigger + poll
    ContinuousModeSweep     (BaseVNASweep)  -- streaming callback loop
    VNAGUIModeSweepTest     (SingleModeSweep, ContinuousModeSweep)
                            -- dispatches to the correct parent based on mode

SCPI commands used -- all documented in ProgrammingGuide.pdf
------------------------------------------------------------
  *IDN?                        4.1.1   identification string
  DEVice:CONNect?              4.2.2   serial of connected device
  DEVice:MODE VNA              4.2.6   switch to VNA mode
  VNA:SWEEP FREQUENCY          4.3.1   frequency-sweep type
  VNA:STIMulus:LVL    <dBm>    4.3.24  stimulus output power
  VNA:ACquisition:IFBW <Hz>    4.3.13  IF bandwidth
  VNA:ACquisition:AVG  <n>     4.3.16  averaging count
  VNA:ACquisition:POINTS <n>   4.3.15  points per sweep
  VNA:FREQuency:START <Hz>     4.3.3   start frequency
  VNA:FREQuency:STOP  <Hz>     4.3.5   stop frequency
  VNA:ACquisition:FINished?    4.3.18  TRUE when averaging complete
  VNA:ACquisition:SINGLE <B>   4.3.20  TRUE=single / FALSE=continuous
  VNA:ACquisition:STOP         4.3.12  halt acquisition
  VNA:ACquisition:RUN          4.3.11  start acquisition
  VNA:TRACe:DATA? S11          4.3.27  [freq,re,im] tuples
  VNA:CALibration:LOAD? <file> 4.3.55  load cal file (query)
  VNA:CALibration:ACTIVE?      4.3.45  active cal type (query)

Prerequisites
-------------
CONTINUOUS MODE:
  * The VNA Calibrated Data streaming server (port 19001) is automatically
    enabled by the script if not already running.  The first continuous-mode
    run will restart the GUI to enable streaming; subsequent runs will use
    the fast path (no restart needed).

BOTH MODES:
  * A calibrated 50-ohm matched load is connected to port 1.
  * The data/ directory is a sibling of scripts/ under LibreVNA-dev/.

Usage
-----
    uv run python 6_librevna_gui_mode_sweep_test.py                  # single, default config
    uv run python 6_librevna_gui_mode_sweep_test.py --mode continuous
    uv run python 6_librevna_gui_mode_sweep_test.py --config /path/to/other.yaml --no-save
"""

import sys
import os
import platform
import math
import time
import socket
import subprocess
import threading
import importlib
import argparse
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from datetime import datetime

import yaml
import numpy as np
from prettytable import PrettyTable
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ---------------------------------------------------------------------------
# Paths -- all relative to this script's location so the whole tree is
# portable regardless of where it is checked out.
# ---------------------------------------------------------------------------

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Insert SCRIPT_DIR so that "from libreVNA import libreVNA" resolves to the
# co-located wrapper regardless of cwd.
sys.path.insert(0, SCRIPT_DIR)
from libreVNA import libreVNA  # noqa: E402

# ---------------------------------------------------------------------------
# Import load_calibration from script 2 via importlib.
# Script 2's filename begins with a digit, which is not a valid Python
# identifier, so a normal "import" statement cannot be used.
# ---------------------------------------------------------------------------
_mod2 = importlib.import_module("2_s11_cal_verification_sweep")
load_calibration = _mod2.load_calibration

# ---------------------------------------------------------------------------
# Module-level constants (GUI lifecycle, SCPI connection)
# ---------------------------------------------------------------------------

# OS-dependent GUI binary path
if platform.system() == "Windows":
    GUI_BINARY = os.path.normpath(
        os.path.join(SCRIPT_DIR, "..", "tools", "LibreVNA-GUI", "release", "LibreVNA-GUI.exe")
    )
else:
    GUI_BINARY = os.path.normpath(
        os.path.join(SCRIPT_DIR, "..", "tools", "LibreVNA-GUI")
    )
CAL_FILE_PATH = os.path.normpath(
    os.path.join(SCRIPT_DIR, "..", "calibration", "SOLT_1_2_43G-2_45G_300pt.cal")
)
DATA_DIR = os.path.normpath(os.path.join(SCRIPT_DIR, "..", "data"))

SCPI_HOST = "localhost"
SCPI_PORT = 1234
GUI_START_TIMEOUT_S = 30.0

# Polling / timeout constants for the single-sweep poll loop
POLL_INTERVAL_S = 0.01
SWEEP_TIMEOUT_S = 60.0

# Streaming port for continuous mode (VNA Calibrated Data)
STREAMING_PORT = 19001
CONTINUOUS_TIMEOUT_S = 300  # hard ceiling on event.wait()

# ---------------------------------------------------------------------------
# Console-output helpers (identical style to scripts 1-5)
# ---------------------------------------------------------------------------


def _section(title):
    """Print a dashed section header."""
    width = 70
    print("\n" + "=" * width)
    print("  " + title)
    print("=" * width)


def _subsection(title):
    """Print a lighter sub-header."""
    print("\n  --- " + title + " ---")


# ---------------------------------------------------------------------------
# SweepResult dataclass
# ---------------------------------------------------------------------------


@dataclass
class SweepResult:
    """Container for one complete IFBW run (all sweeps at that setting)."""

    mode: str  # "single" or "continuous"
    ifbw_hz: int
    sweep_times: list  # list[float] wall-clock seconds per sweep
    all_s11_db: list  # list[list[float]] [sweep_idx][point_idx]
    freq_hz: list  # list[float] frequency axis
    noise_floor: float = 0.0  # filled by compute_metrics
    trace_jitter: float = 0.0  # filled by compute_metrics


# ===========================================================================
# BaseVNASweep  --  abstract base class
# ===========================================================================


class BaseVNASweep(ABC):
    """
    Shared lifecycle, GUI management, calibration loading, metric
    computation, console summary, and xlsx export.  Concrete sweep
    logic (configure + run) is left to the subclasses.
    """

    def __init__(self, config_path, mode, summary=True, save_data=True):
        """
        Parameters
        ----------
        config_path : str   -- absolute path to the YAML config file.
        mode        : str   -- "single" or "continuous".
        summary     : bool  -- if True, print the PrettyTable at the end.
        save_data   : bool  -- if True, write the xlsx workbook.
        """
        self.mode = mode
        self.summary = summary
        self.save_data = save_data

        # -- Load YAML config --------------------------------------------------
        with open(config_path, "r") as fh:
            raw = yaml.safe_load(fh)

        cfg = raw["configurations"]
        tgt = raw["target"]

        self.start_freq_hz = int(cfg["start_frequency"])
        self.stop_freq_hz = int(cfg["stop_frequency"])
        self.num_points = int(cfg["num_points"])
        self.stim_lvl_dbm = int(cfg["stim_lvl_dbm"])
        self.avg_count = int(cfg["avg_count"])
        self.num_sweeps = int(cfg["num_sweeps"])

        # Normalise ifbw_values: accept a single int OR a list of ints.
        raw_ifbw = tgt["ifbw_values"]
        if isinstance(raw_ifbw, list):
            self.ifbw_values = [int(v) for v in raw_ifbw]
        else:
            self.ifbw_values = [int(raw_ifbw)]

    # -----------------------------------------------------------------------
    # GUI lifecycle
    # -----------------------------------------------------------------------

    def start_gui(self):
        """
        Launch LibreVNA-GUI in headless mode and poll TCP port SCPI_PORT
        until the SCPI server accepts a connection.  Returns the Popen handle.
        """
        _section("STARTING LibreVNA-GUI")

        env = os.environ.copy()
        env["QT_QPA_PLATFORM"] = "offscreen"

        proc = subprocess.Popen(
            [GUI_BINARY, "--port", str(SCPI_PORT)],
            env=env,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        print("  GUI PID         : {}".format(proc.pid))

        deadline = time.time() + GUI_START_TIMEOUT_S
        while True:
            try:
                s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                s.settimeout(0.5)
                s.connect((SCPI_HOST, SCPI_PORT))
                s.close()
                print("  SCPI server     : ready on {}:{}".format(SCPI_HOST, SCPI_PORT))
                return proc
            except (ConnectionRefusedError, OSError):
                s.close()
            if time.time() > deadline:
                proc.terminate()
                proc.wait()
                raise RuntimeError(
                    "LibreVNA-GUI did not open SCPI server on port {} "
                    "within {:.0f} s".format(SCPI_PORT, GUI_START_TIMEOUT_S)
                )
            time.sleep(0.25)

    def stop_gui(self, proc):
        """
        Terminate the GUI subprocess gracefully (SIGTERM) with a 5 s
        timeout, escalating to SIGKILL if needed.
        """
        _section("STOPPING LibreVNA-GUI")
        proc.terminate()
        try:
            proc.wait(timeout=5)
            print("  GUI terminated  : PID {} (clean)".format(proc.pid))
        except subprocess.TimeoutExpired:
            proc.kill()
            proc.wait()
            print("  GUI killed      : PID {} (forced)".format(proc.pid))

    # -----------------------------------------------------------------------
    # Device connection
    # -----------------------------------------------------------------------

    def connect_and_verify(self):
        """
        Open a TCP connection to LibreVNA-GUI and confirm a hardware device
        is attached.  Raises RuntimeError on any failure (the GUI was just
        started by this script so a connection problem is unexpected).

        Returns
        -------
        libreVNA
            Connected and verified wrapper instance.
        """
        _section("DEVICE CONNECTION")

        try:
            vna = libreVNA(host=SCPI_HOST, port=SCPI_PORT)
            print("  TCP connection  : OK  ({}:{})".format(SCPI_HOST, SCPI_PORT))
        except Exception as exc:
            raise RuntimeError(
                "Could not connect to LibreVNA-GUI at {}:{}: {}".format(
                    SCPI_HOST, SCPI_PORT, exc
                )
            )

        _subsection("*IDN? identification")
        try:
            idn_raw = vna.query("*IDN?")
            print("  Raw response    : {}".format(idn_raw))
            parts = [p.strip() for p in idn_raw.split(",")]
            labels = ["Manufacturer", "Model", "Serial (IDN)", "SW Version"]
            for label, val in zip(labels, parts):
                print("    {:<22s}: {}".format(label, val))
        except Exception as exc:
            print("  [WARN] *IDN? query failed: {}".format(exc))

        _subsection("DEVice:CONNect? -- device serial")
        try:
            dev_serial = vna.query(":DEV:CONN?")
            print("  Live serial     : {}".format(dev_serial))
            if dev_serial == "Not connected":
                raise RuntimeError(
                    "LibreVNA-GUI is not connected to any hardware device."
                )
        except RuntimeError:
            raise
        except Exception as exc:
            raise RuntimeError("DEVice:CONNect? query failed: {}".format(exc))

        return vna

    # -----------------------------------------------------------------------
    # Calibration
    # -----------------------------------------------------------------------

    def load_calibration(self, vna):
        """
        Delegate to the load_calibration() imported from script 2.

        Parameters
        ----------
        vna : libreVNA
            Connected wrapper instance.
        """
        load_calibration(vna)


    def enable_streaming_server(self, vna):
        """
        Ensure the VNA Calibrated Data streaming server is enabled on port 19001.

        If the streaming server is already enabled (port 19001 is listening),
        this method returns False immediately (fast path).

        If the streaming server is disabled, this method enables it via SCPI:
            :DEV:PREF StreamingServers.VNACalibratedData.enabled true
            :DEV:APPLYPREFERENCES

        Note that APPLYPREFERENCES saves the preference to disk and terminates
        the GUI.  The caller (run()) must stop the old GUI and start a new one.

        Parameters
        ----------
        vna : libreVNA
            The currently active connection (will become stale after this call).

        Returns
        -------
        bool
            True if the preference was set (caller must restart GUI).
            False if streaming was already enabled (no restart needed).
        """
        _section("STREAMING SERVER SETUP")

        # Test if streaming server is already listening
        print("  Testing port {} ...".format(STREAMING_PORT))
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.settimeout(1.0)
            s.connect((SCPI_HOST, STREAMING_PORT))
            s.close()
            print("  Status          : already enabled (fast path)")
            return False  # no restart needed
        except (ConnectionRefusedError, OSError):
            s.close()
            print("  Status          : not enabled, enabling now...")

        # Enable streaming server via SCPI
        # DEV:PREF set returns CME bit even on success -- use check=False
        print("  Sending         : DEV:PREF StreamingServers.VNACalibratedData.enabled true")
        vna.cmd(":DEV:PREF StreamingServers.VNACalibratedData.enabled true", check=False)

        print("  Sending         : DEV:APPLYPREFERENCES")
        vna.cmd(":DEV:APPLYPREFERENCES", check=False)

        # APPLYPREFERENCES saves the preference to disk and terminates the GUI.
        # Give it a moment to save, then let the caller handle the restart.
        print("  Preference      : saved to disk")
        print("  [INFO] GUI will terminate now; caller must restart it")
        time.sleep(2)  # brief pause to ensure preference is saved

        return True  # caller must restart GUI

    # -----------------------------------------------------------------------
    # Abstract methods -- implemented by the single / continuous subclasses
    # -----------------------------------------------------------------------

    @abstractmethod
    def configure_sweep(self, vna, ifbw_hz):
        """Configure the VNA for a sweep at the given IFBW."""
        ...

    @abstractmethod
    def run_sweeps(self, vna, ifbw_hz):
        """
        Execute self.num_sweeps sweeps and return a SweepResult.

        Parameters
        ----------
        vna     : libreVNA
        ifbw_hz : int

        Returns
        -------
        SweepResult
        """
        ...

    # -----------------------------------------------------------------------
    # Metric computation (shared)
    # -----------------------------------------------------------------------

    def compute_metrics(self, result):
        """
        Fill result.noise_floor and result.trace_jitter in-place.

        Noise floor  = mean( per-sweep mean S11_dB across all freq points )
        Trace jitter = mean( per-point std across sweeps, ddof=1 )

        Parameters
        ----------
        result : SweepResult
            Must have all_s11_db populated.
        """
        s11_arr = np.array(result.all_s11_db)  # shape (num_sweeps, num_points)

        # Noise floor
        per_sweep_means = np.mean(s11_arr, axis=1)  # shape (num_sweeps,)
        result.noise_floor = float(np.mean(per_sweep_means))

        # Trace jitter
        per_point_stds = np.std(s11_arr, axis=0, ddof=1)  # shape (num_points,)
        result.trace_jitter = float(np.mean(per_point_stds))

    # -----------------------------------------------------------------------
    # Console summary (PrettyTable)
    # -----------------------------------------------------------------------

    def print_summary(self, all_results):
        """
        Print a PrettyTable with one row per IFBW result.

        Columns: Mode | IFBW kHz | Mean Time s | Std s | Min s | Max s |
                 Rate Hz | Noise Floor dB | Jitter dB

        Parameters
        ----------
        all_results : list[SweepResult]
        """
        _section("SWEEP TEST SUMMARY  ({} mode)".format(self.mode))

        table = PrettyTable()
        table.field_names = [
            "Mode",
            "IFBW (kHz)",
            "Mean Time (s)",
            "Std Dev (s)",
            "Min Time (s)",
            "Max Time (s)",
            "Rate (Hz)",
            "Noise Floor (dB)",
            "Trace Jitter (dB)",
        ]

        for r in all_results:
            times_arr = np.array(r.sweep_times)
            mean_t = float(np.mean(times_arr))
            std_t = float(np.std(times_arr, ddof=1)) if len(times_arr) > 1 else 0.0
            min_t = float(np.min(times_arr))
            max_t = float(np.max(times_arr))
            rate = 1.0 / mean_t if mean_t > 0 else float("inf")

            table.add_row(
                [
                    r.mode,
                    "{:d}".format(r.ifbw_hz // 1000),
                    "{:.4f}".format(mean_t),
                    "{:.4f}".format(std_t),
                    "{:.4f}".format(min_t),
                    "{:.4f}".format(max_t),
                    "{:.2f}".format(rate),
                    "{:.2f}".format(r.noise_floor),
                    "{:.4f}".format(r.trace_jitter),
                ]
            )

        print(table)

    # -----------------------------------------------------------------------
    # xlsx export
    # -----------------------------------------------------------------------

    def save_xlsx(self, all_results):
        """
        Write a multi-sheet xlsx workbook.

        Sheet layout
        ------------
        "Summary"           -- one row per IFBW with all metrics.
        "IFBW_{n}kHz"       -- per-IFBW detail: config, timing, S11 traces,
                               and metrics blocks.

        Parameters
        ----------
        all_results : list[SweepResult]

        Returns
        -------
        str
            Absolute path of the written xlsx file.
        """
        # -- Output path: ../data/YYYYMMDD/ -----------------------------------
        today = datetime.now().strftime("%Y%m%d")
        out_dir = os.path.join(DATA_DIR, today)
        os.makedirs(out_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "{}_sweep_test_{}.xlsx".format(self.mode, timestamp)
        full_path = os.path.join(out_dir, filename)

        # -- Style definitions ------------------------------------------------
        bold_font = Font(bold=True)
        title_font = Font(bold=True, size=13)
        section_font = Font(bold=True, size=12)
        header_fill = PatternFill(
            start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"
        )

        wb = Workbook()

        # ---- Sheet: Summary -------------------------------------------------
        ws = wb.active
        ws.title = "Summary"

        # Row 1: merged title
        ws.cell(
            row=1, column=1, value="VNA Sweep Test Summary -- {} mode".format(self.mode)
        )
        ws.cell(row=1, column=1).font = title_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

        # Row 2: blank (already empty)

        # Row 3: headers
        summary_headers = [
            "Mode",
            "IFBW (kHz)",
            "Mean Time (s)",
            "Mean Time (ms)",
            "Std Dev (s)",
            "Min Time (s)",
            "Max Time (s)",
            "Rate (Hz)",
            "Noise Floor (dB)",
            "Trace Jitter (dB)",
        ]
        for col_idx, hdr in enumerate(summary_headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=hdr)
            cell.font = bold_font
            cell.fill = header_fill

        # Row 4+: data
        for row_offset, r in enumerate(all_results):
            row = 4 + row_offset
            times_arr = np.array(r.sweep_times)
            mean_t = float(np.mean(times_arr))
            std_t = float(np.std(times_arr, ddof=1)) if len(times_arr) > 1 else 0.0
            min_t = float(np.min(times_arr))
            max_t = float(np.max(times_arr))
            rate = 1.0 / mean_t if mean_t > 0 else float("inf")

            values = [
                r.mode,
                r.ifbw_hz // 1000,
                round(mean_t, 4),
                round(mean_t * 1000, 4),
                round(std_t, 4),
                round(min_t, 4),
                round(max_t, 4),
                round(rate, 2),
                round(r.noise_floor, 2),
                round(r.trace_jitter, 2),
            ]
            for col_idx, val in enumerate(values, start=1):
                ws.cell(row=row, column=col_idx, value=val)

        # ---- Per-IFBW sheets ------------------------------------------------
        for r in all_results:
            ifbw_khz = r.ifbw_hz // 1000
            sheet_name = "IFBW_{}kHz".format(ifbw_khz)
            ws = wb.create_sheet(title=sheet_name)

            # ==== Config block (rows 1-9) ====================================
            ws.cell(row=1, column=1, value="Configuration")
            ws.cell(row=1, column=1).font = section_font

            config_rows = [
                ("Mode", self.mode),
                ("IFBW (kHz)", ifbw_khz),
                ("Start Freq (Hz)", self.start_freq_hz),
                ("Stop Freq (Hz)", self.stop_freq_hz),
                ("Points", self.num_points),
                ("STIM Level (dBm)", self.stim_lvl_dbm),
                ("Avg Count", self.avg_count),
                ("Num Sweeps", self.num_sweeps),
            ]
            for i, (label, value) in enumerate(config_rows):
                row = 2 + i  # rows 2-9
                ws.cell(row=row, column=1, value=label).font = bold_font
                ws.cell(row=row, column=2, value=value)
            # Row 9 is the last config row (2 + 7); row 10 is blank (gap already)

            # ==== Timing block (starts row 10) ===============================
            timing_header_row = 10
            ws.cell(row=timing_header_row, column=1, value="Timing")
            ws.cell(row=timing_header_row, column=1).font = section_font

            timing_cols = [
                "Sweep #",
                "Sweep Time (s)",
                "Sweep Time (ms)",
                "Update Rate (Hz)",
            ]
            col_header_row = timing_header_row + 1  # row 11
            for col_idx, hdr in enumerate(timing_cols, start=1):
                cell = ws.cell(row=col_header_row, column=col_idx, value=hdr)
                cell.font = bold_font
                cell.fill = header_fill

            for sweep_idx, t in enumerate(r.sweep_times):
                row = col_header_row + 1 + sweep_idx  # row 12+
                rate = 1.0 / t if t > 0 else float("inf")
                ws.cell(row=row, column=1, value=sweep_idx + 1)
                ws.cell(row=row, column=2, value=round(t, 4))
                ws.cell(row=row, column=3, value=round(t * 1000, 4))
                ws.cell(row=row, column=4, value=round(rate, 2))

            # blank row after timing
            traces_section_row = col_header_row + 1 + len(r.sweep_times) + 1

            # ==== S11 Traces block ============================================
            ws.cell(row=traces_section_row, column=1, value="S11 Traces")
            ws.cell(row=traces_section_row, column=1).font = section_font

            trace_col_header_row = traces_section_row + 1
            # Column headers: Frequency (Hz) | Sweep_1 S11 (dB) | ...
            ws.cell(row=trace_col_header_row, column=1, value="Frequency (Hz)")
            ws.cell(row=trace_col_header_row, column=1).font = bold_font
            ws.cell(row=trace_col_header_row, column=1).fill = header_fill
            for s_idx in range(len(r.all_s11_db)):
                col = 2 + s_idx
                ws.cell(
                    row=trace_col_header_row,
                    column=col,
                    value="Sweep_{} S11 (dB)".format(s_idx + 1),
                )
                ws.cell(row=trace_col_header_row, column=col).font = bold_font
                ws.cell(row=trace_col_header_row, column=col).fill = header_fill

            # Data rows: one per frequency point
            for pt_idx, freq in enumerate(r.freq_hz):
                row = trace_col_header_row + 1 + pt_idx
                ws.cell(row=row, column=1, value=freq)
                for s_idx in range(len(r.all_s11_db)):
                    ws.cell(
                        row=row,
                        column=2 + s_idx,
                        value=round(r.all_s11_db[s_idx][pt_idx], 4),
                    )

            # blank row after traces
            metrics_section_row = trace_col_header_row + 1 + len(r.freq_hz) + 1

            # ==== Metrics block ===============================================
            ws.cell(row=metrics_section_row, column=1, value="Metrics")
            ws.cell(row=metrics_section_row, column=1).font = section_font

            ws.cell(row=metrics_section_row + 1, column=1, value="Noise Floor (dB)")
            ws.cell(row=metrics_section_row + 1, column=1).font = bold_font
            ws.cell(
                row=metrics_section_row + 1, column=2, value=round(r.noise_floor, 4)
            )

            ws.cell(row=metrics_section_row + 2, column=1, value="Trace Jitter (dB)")
            ws.cell(row=metrics_section_row + 2, column=1).font = bold_font
            ws.cell(
                row=metrics_section_row + 2, column=2, value=round(r.trace_jitter, 6)
            )

        # -- Write workbook ---------------------------------------------------
        wb.save(full_path)
        return full_path

    # -----------------------------------------------------------------------
    # Lifecycle hooks  --  override in subclasses for one-time setup / teardown
    # -----------------------------------------------------------------------

    def pre_loop_reset(self, vna):
        """Called once before the IFBW loop.  Override for mode-specific one-time setup."""
        pass

    def post_loop_teardown(self, vna):
        """Called once after the IFBW loop completes.  Override for cleanup."""
        pass

    # -----------------------------------------------------------------------
    # Main orchestration
    # -----------------------------------------------------------------------

    def run(self):
        """
        Top-level entry point.

        Sequence
        --------
        1.  start_gui()
        2.  connect_and_verify()
        3.  enable_streaming_server() (continuous mode only; auto-enables + may restart GUI)
        4.  load_calibration()
        5.  pre_loop_reset()   (one-time mode-specific setup)
        6.  For each IFBW: configure_sweep() -> run_sweeps() -> compute_metrics()
        7.  post_loop_teardown()   (one-time mode-specific cleanup)
        8.  print_summary()  (if self.summary)
        9.  save_xlsx()      (if self.save_data)
        10. stop_gui()       (in finally block)
        """
        gui_proc = self.start_gui()
        try:
            vna = self.connect_and_verify()

            # If continuous mode, enable streaming server (may restart GUI)
            if self.mode == "continuous":
                needs_restart = self.enable_streaming_server(vna)
                if needs_restart:
                    # APPLYPREFERENCES terminated the old GUI; restart it
                    _section("RESTARTING GUI")
                    self.stop_gui(gui_proc)
                    gui_proc = self.start_gui()
                    vna = self.connect_and_verify()

            self.load_calibration(vna)

            all_results = []

            self.pre_loop_reset(vna)

            for ifbw_hz in self.ifbw_values:
                _section(
                    "IFBW = {} kHz  --  {} mode".format(ifbw_hz // 1000, self.mode)
                )

                self.configure_sweep(vna, ifbw_hz)
                result = self.run_sweeps(vna, ifbw_hz)
                self.compute_metrics(result)

                # Per-IFBW progress
                _subsection("IFBW {} kHz -- metrics".format(ifbw_hz // 1000))
                times_arr = np.array(result.sweep_times)
                print(
                    "    Mean sweep time : {:.4f} s".format(float(np.mean(times_arr)))
                )
                print(
                    "    Update rate     : {:.2f} Hz".format(
                        1.0 / float(np.mean(times_arr))
                        if float(np.mean(times_arr)) > 0
                        else float("inf")
                    )
                )
                print("    Noise floor     : {:.2f} dB".format(result.noise_floor))
                print("    Trace jitter    : {:.4f} dB".format(result.trace_jitter))

                all_results.append(result)

            self.post_loop_teardown(vna)

            # -- Console summary ----------------------------------------------
            if self.summary:
                self.print_summary(all_results)

            # -- xlsx export --------------------------------------------------
            if self.save_data:
                _section("SAVING RESULTS")
                xlsx_path = self.save_xlsx(all_results)
                print("  xlsx written    : {}".format(xlsx_path))

            print()  # trailing blank line

        finally:
            self.stop_gui(gui_proc)


# ===========================================================================
# SingleModeSweep
# ===========================================================================


class SingleModeSweep(BaseVNASweep):
    """
    Single-sweep mode: trigger via FREQuency:STOP, poll FIN?, read trace.

    configure_sweep() sets all sweep parameters EXCEPT FREQuency:STOP.
    Each sweep in the loop is triggered by sending FREQuency:STOP, which
    both sets the stop frequency and initiates acquisition.  This mirrors
    the proven trigger-and-poll protocol used in script 4.
    """

    def configure_sweep(self, vna, ifbw_hz):
        """
        Single-mode configuration -- everything except STOP.

        SCPI sequence
        -------------
        :DEV:MODE VNA
        :VNA:SWEEP FREQUENCY
        :VNA:STIM:LVL <dBm>
        :VNA:ACQ:IFBW  <Hz>
        :VNA:ACQ:AVG   <n>
        :VNA:ACQ:POINTS <n>
        :VNA:FREQuency:START <Hz>
        (STOP is the acquisition trigger, sent per-sweep in the loop)
        """
        _subsection(
            "Single-mode configuration  (IFBW = {} kHz)".format(ifbw_hz // 1000)
        )

        vna.cmd(":DEV:MODE VNA")
        print("  Mode            : VNA")

        vna.cmd(":VNA:SWEEP FREQUENCY")
        print("  Sweep type      : FREQUENCY")

        vna.cmd(":VNA:STIM:LVL {}".format(self.stim_lvl_dbm))
        print("  Stimulus level  : {} dBm".format(self.stim_lvl_dbm))

        vna.cmd(":VNA:ACQ:IFBW {}".format(ifbw_hz))
        print("  IF bandwidth    : {} Hz  ({} kHz)".format(ifbw_hz, ifbw_hz // 1000))

        vna.cmd(":VNA:ACQ:AVG {}".format(self.avg_count))
        print("  Averaging       : {} sweep(s)".format(self.avg_count))

        vna.cmd(":VNA:ACQ:POINTS {}".format(self.num_points))
        print("  Points          : {}".format(self.num_points))

        vna.cmd(":VNA:FREQuency:START {}".format(self.start_freq_hz))
        print(
            "  Start freq      : {} Hz  ({:.3f} GHz)".format(
                self.start_freq_hz, self.start_freq_hz / 1e9
            )
        )

        print("  Stop freq       : (will be sent as sweep trigger)")

    def run_sweeps(self, vna, ifbw_hz):
        """Dispatch to the timed single-sweep loop."""
        return self._single_sweep_loop(vna, ifbw_hz)

    def _single_sweep_loop(self, vna, ifbw_hz):
        """
        Execute self.num_sweeps consecutive S11 sweeps using the
        trigger-poll-read protocol from scripts 3 and 4.

        Timing protocol per iteration
        -----------------------------
        1. vna.cmd(":VNA:FREQuency:STOP <Hz>")   -- triggers acquisition
        2. t_start = time.time()
        3. Poll :VNA:ACQ:FIN? every POLL_INTERVAL_S until "TRUE"
        4. t_end = time.time()   -- sweep_time = t_end - t_start
        5. Read :VNA:TRACE:DATA? S11             -- OUTSIDE timed window
        6. Parse + convert to dB                 -- OUTSIDE timed window

        Returns
        -------
        SweepResult
        """
        _subsection("Single-sweep loop  ({} sweeps)".format(self.num_sweeps))

        sweep_times = []
        all_s11_db = []
        freq_hz_axis = []

        for i in range(self.num_sweeps):

            # -- Trigger via FREQuency:STOP (sets endpoint + starts sweep) --------
            vna.cmd(":VNA:FREQuency:STOP {}".format(self.stop_freq_hz))

            # -- Time: start ----------------------------------------------------
            t_start = time.time()

            # -- Poll for completion --------------------------------------------
            while True:
                finished = vna.query(":VNA:ACQ:FIN?")
                if finished == "TRUE":
                    break
                if time.time() - t_start > SWEEP_TIMEOUT_S:
                    raise TimeoutError(
                        "IFBW={} kHz, sweep {}/{}: VNA:ACQ:FIN? did not return "
                        "TRUE within {:.0f} s (last response: '{}')".format(
                            ifbw_hz // 1000,
                            i + 1,
                            self.num_sweeps,
                            SWEEP_TIMEOUT_S,
                            finished,
                        )
                    )
                time.sleep(POLL_INTERVAL_S)

            # -- Time: end ------------------------------------------------------
            t_end = time.time()
            sweep_time = t_end - t_start
            sweep_times.append(sweep_time)

            # -- Read trace (outside timed window) ------------------------------
            raw_data = vna.query(":VNA:TRACE:DATA? S11")
            trace = libreVNA.parse_VNA_trace_data(raw_data)

            # -- Convert to dB --------------------------------------------------
            sweep_freq = []
            sweep_s11db = []
            for fq, gamma in trace:
                magnitude = abs(gamma)
                if magnitude < 1e-12:
                    magnitude = 1e-12
                s11_db = 20.0 * math.log10(magnitude)
                sweep_freq.append(float(fq))
                sweep_s11db.append(float(s11_db))

            if i == 0:
                freq_hz_axis = sweep_freq

            all_s11_db.append(sweep_s11db)

            # -- Progress line --------------------------------------------------
            update_rate = 1.0 / sweep_time if sweep_time > 0 else float("inf")
            print(
                "    Sweep {:>2d}/{:<2d}  :  {:.4f} s  ({:.1f} Hz)".format(
                    i + 1, self.num_sweeps, sweep_time, update_rate
                )
            )

        return SweepResult(
            mode="single",
            ifbw_hz=ifbw_hz,
            sweep_times=sweep_times,
            all_s11_db=all_s11_db,
            freq_hz=freq_hz_axis,
        )


# ===========================================================================
# ContinuousModeSweep
# ===========================================================================


class ContinuousModeSweep(BaseVNASweep):
    """
    Continuous-sweep mode: streaming callback on port 19001.

    The streaming callback is registered ONCE in pre_loop_reset() and
    removed ONCE in post_loop_teardown().  Between IFBWs the callback
    stays connected; only the mutable state it reads is swapped.  This
    avoids the remove + re-add path that triggers a known bug in
    libreVNA.py line 148 (the thread-exit guard checks the wrong list
    length).

    configure_sweep sends the full configuration INCLUDING STOP (no
    ACQ:STOP / SINGLE TRUE prefix -- those are handled by
    pre_loop_reset / post_loop_teardown).
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._state_holder = [None]  # mutable container; callback always reads [0]
        self._stream_callback = None  # the single registered callback closure

    # -- Inner helper: shared mutable state guarded by a Lock -----------------

    class _SweepState:
        """
        All mutable state touched by the streaming callback, bundled into a
        single object so that the lock scope is unambiguous.
        """

        def __init__(self, num_points, num_sweeps):
            self.lock = threading.Lock()
            self.done_event = threading.Event()
            self.num_points = num_points
            self.num_sweeps = num_sweeps
            self.sweep_count = 0
            self.sweep_start_time = 0.0
            self.sweep_end_times = []  # list[float]
            self.sweep_start_times = []  # list[float]
            self.current_s11 = []  # list[complex], current sweep
            self.all_s11_complex = []  # list[list[complex]], all completed sweeps

    # -----------------------------------------------------------------------

    def configure_sweep(self, vna, ifbw_hz):
        """
        Continuous-mode configuration -- full sequence including STOP.

        SCPI sequence
        -------------
        :DEV:MODE VNA
        :VNA:SWEEP FREQUENCY
        :VNA:STIM:LVL <dBm>
        :VNA:ACQ:IFBW  <Hz>
        :VNA:ACQ:AVG   <n>
        :VNA:ACQ:POINTS <n>
        :VNA:FREQuency:START <Hz>
        :VNA:FREQuency:STOP  <Hz>   -- included because STOP is NOT the trigger here
        """
        _subsection(
            "Continuous-mode configuration  (IFBW = {} kHz)".format(ifbw_hz // 1000)
        )

        vna.cmd(":DEV:MODE VNA")
        print("  Mode            : VNA")

        vna.cmd(":VNA:SWEEP FREQUENCY")
        print("  Sweep type      : FREQUENCY")

        vna.cmd(":VNA:STIM:LVL {}".format(self.stim_lvl_dbm))
        print("  Stimulus level  : {} dBm".format(self.stim_lvl_dbm))

        vna.cmd(":VNA:ACQ:IFBW {}".format(ifbw_hz))
        print("  IF bandwidth    : {} Hz  ({} kHz)".format(ifbw_hz, ifbw_hz // 1000))

        vna.cmd(":VNA:ACQ:AVG {}".format(self.avg_count))
        print("  Averaging       : {} sweep(s)".format(self.avg_count))

        vna.cmd(":VNA:ACQ:POINTS {}".format(self.num_points))
        print("  Points          : {}".format(self.num_points))

        vna.cmd(":VNA:FREQuency:START {}".format(self.start_freq_hz))
        print(
            "  Start freq      : {} Hz  ({:.3f} GHz)".format(
                self.start_freq_hz, self.start_freq_hz / 1e9
            )
        )

        vna.cmd(":VNA:FREQuency:STOP {}".format(self.stop_freq_hz))
        print(
            "  Stop freq       : {} Hz  ({:.3f} GHz)".format(
                self.stop_freq_hz, self.stop_freq_hz / 1e9
            )
        )

    def run_sweeps(self, vna, ifbw_hz):
        """Dispatch to the streaming callback loop."""
        return self._continuous_sweep_loop(vna, ifbw_hz)

    def _make_callback(self, state_holder):
        """
        Return a closure that reads its state from state_holder[0].

        state_holder is a one-element list.  The main thread swaps [0] to a
        fresh _SweepState between IFBWs; the callback picks up the new state
        on the next point without any remove/re-add of the TCP connection.
        """

        def _callback(data):
            state = state_holder[0]
            if state is None:
                return  # gap during state swap -- drop silently
            if "Z0" not in data:
                return

            point_num = data["pointNum"]
            s11_complex = data["measurements"].get("S11", complex(0, 0))

            with state.lock:
                if point_num == 0:
                    state.sweep_start_time = time.time()
                    state.current_s11 = []

                state.current_s11.append(s11_complex)

                if point_num == state.num_points - 1:
                    sweep_end = time.time()
                    state.sweep_end_times.append(sweep_end)
                    state.sweep_start_times.append(state.sweep_start_time)
                    state.all_s11_complex.append(list(state.current_s11))
                    state.sweep_count += 1

                    if state.sweep_count >= state.num_sweeps:
                        state.done_event.set()

        return _callback

    def pre_loop_reset(self, vna):
        """
        Register streaming callback ONCE and prepare continuous mode.
        
        The streaming server is guaranteed to be enabled by run() before
        this method is called, so we can directly connect without error
        handling.
        """
        _subsection("Continuous-mode streaming setup (one-time)")
        
        # Create the persistent callback referencing the mutable state holder
        self._stream_callback = self._make_callback(self._state_holder)
        
        # Stop any residual acquisition before connecting
        vna.cmd(":VNA:ACQ:STOP")
        print("  Pre-stop        : sent")
        
        # Register the streaming callback (streaming server is guaranteed enabled)
        vna.add_live_callback(STREAMING_PORT, self._stream_callback)
        print("  Streaming       : callback registered on port {}".format(STREAMING_PORT))

    def post_loop_teardown(self, vna):
        """Stop acquisition, restore single mode, remove the streaming callback."""
        _subsection("Continuous-mode streaming teardown")

        vna.cmd(":VNA:ACQ:STOP")
        print("  Acquisition     : stopped")

        vna.cmd(":VNA:ACQ:SINGLE TRUE")
        print("  Sweep mode      : restored to SINGLE (TRUE)")

        if self._stream_callback is not None:
            vna.remove_live_callback(STREAMING_PORT, self._stream_callback)
            self._stream_callback = None
            print("  Streaming       : callback removed")

    def _continuous_sweep_loop(self, vna, ifbw_hz):
        """
        Run self.num_sweeps continuous sweeps for one IFBW value.

        The streaming callback is already registered (by pre_loop_reset).
        This method only manages per-IFBW acquisition start/stop and state.

        Sequence per IFBW
        ------------------
        1.  ACQ:STOP              -- halt any sweeps from the previous IFBW
        2.  sleep 0.1 s           -- drain buffered streaming points so the
                                     fresh state does not see stale data
        3.  swap state_holder[0]  -- atomically switch the callback to a
                                     new _SweepState (GIL makes list-item
                                     assignment atomic)
        4.  ACQ:SINGLE FALSE      -- ensure continuous mode (idempotent)
        5.  ACQ:RUN               -- start back-to-back sweeps
        6.  wait on done_event    -- callback sets it when sweep_count reaches target
        7.  ACQ:STOP              -- halt after target is reached

        Returns
        -------
        SweepResult
        """
        _subsection(
            "Continuous-sweep loop  ({} sweeps, IFBW {} kHz)".format(
                self.num_sweeps, ifbw_hz // 1000
            )
        )

        # 1. Stop previous IFBW's sweeps
        vna.cmd(":VNA:ACQ:STOP")
        print("  ACQ:STOP        : sent (halt previous IFBW)")

        # 2. Drain any buffered streaming points from the previous sweep
        time.sleep(0.1)

        # 3. Swap in a fresh state -- the callback picks it up on next point
        fresh_state = self._SweepState(self.num_points, self.num_sweeps)
        self._state_holder[0] = fresh_state
        print("  State           : fresh _SweepState installed")

        # 4. Ensure continuous mode (ACQ:STOP does NOT change SINGLE setting,
        #    but belt-and-suspenders is cheap here)
        vna.cmd(":VNA:ACQ:SINGLE FALSE")
        print("  Sweep mode      : CONTINUOUS  (SINGLE FALSE)")

        # 5. Start acquisition
        vna.cmd(":VNA:ACQ:RUN")
        print("  Acquisition     : started (continuous)")
        print(
            "  Collecting {} sweeps via streaming callback ...".format(self.num_sweeps)
        )

        # 6. Wait for completion
        completed = fresh_state.done_event.wait(timeout=CONTINUOUS_TIMEOUT_S)

        if not completed:
            # Tear down before raising
            vna.cmd(":VNA:ACQ:STOP")
            raise TimeoutError(
                "Only {}/{} sweeps received within {} s timeout.  "
                "Check streaming server connectivity.".format(
                    fresh_state.sweep_count, self.num_sweeps, CONTINUOUS_TIMEOUT_S
                )
            )

        # 7. Stop sweeps for this IFBW (do NOT restore SINGLE TRUE here --
        #    that is post_loop_teardown's job)
        vna.cmd(":VNA:ACQ:STOP")
        print("  Acquisition     : stopped")

        # -- Progress printout ---------------------------------------------------
        with fresh_state.lock:
            for i in range(fresh_state.sweep_count):
                duration = (
                    fresh_state.sweep_end_times[i] - fresh_state.sweep_start_times[i]
                )
                dur_rate = 1.0 / duration if duration > 0 else float("inf")
                if i == 0:
                    print(
                        "    Sweep {:>2d}/{:<2d}  :  dur {:.4f} s  ({:.1f} Hz)  "
                        "inter-sweep: --".format(
                            i + 1, fresh_state.sweep_count, duration, dur_rate
                        )
                    )
                else:
                    interval = (
                        fresh_state.sweep_end_times[i]
                        - fresh_state.sweep_end_times[i - 1]
                    )
                    int_rate = 1.0 / interval if interval > 0 else float("inf")
                    print(
                        "    Sweep {:>2d}/{:<2d}  :  dur {:.4f} s  ({:.1f} Hz)  "
                        "inter-sweep: {:.4f} s  ({:.1f} Hz)".format(
                            i + 1,
                            fresh_state.sweep_count,
                            duration,
                            dur_rate,
                            interval,
                            int_rate,
                        )
                    )

        # -- Convert all sweeps to dB -------------------------------------------
        freq_hz = list(
            np.linspace(
                float(self.start_freq_hz), float(self.stop_freq_hz), self.num_points
            )
        )

        all_s11_db = []
        sweep_times = []

        for i in range(fresh_state.sweep_count):
            t = fresh_state.sweep_end_times[i] - fresh_state.sweep_start_times[i]
            sweep_times.append(t)

            raw_s11 = fresh_state.all_s11_complex[i]
            s11_db_list = []
            for gamma in raw_s11:
                magnitude = abs(gamma)
                if magnitude < 1e-12:
                    magnitude = 1e-12
                s11_db_list.append(20.0 * math.log10(magnitude))
            all_s11_db.append(s11_db_list)

        return SweepResult(
            mode="continuous",
            ifbw_hz=ifbw_hz,
            sweep_times=sweep_times,
            all_s11_db=all_s11_db,
            freq_hz=freq_hz,
        )


# ===========================================================================
# VNAGUIModeSweepTest  --  concrete facade
# ===========================================================================


class VNAGUIModeSweepTest(SingleModeSweep, ContinuousModeSweep):
    """
    Dispatching facade.  Inherits from both SingleModeSweep and
    ContinuousModeSweep (MRO: VNAGUIModeSweepTest -> SingleModeSweep ->
    ContinuousModeSweep -> BaseVNASweep).  configure_sweep() and
    run_sweeps() route to the appropriate parent method based on self.mode.
    """

    def __init__(self, config_path, mode="single", summary=True, save_data=True):
        if mode not in ("single", "continuous"):
            raise ValueError(
                "mode must be 'single' or 'continuous', got '{}'".format(mode)
            )
        # BaseVNASweep.__init__ is called once via super() thanks to MRO.
        super().__init__(
            config_path=config_path, mode=mode, summary=summary, save_data=save_data
        )

    def configure_sweep(self, vna, ifbw_hz):
        """Dispatch to SingleModeSweep or ContinuousModeSweep configure."""
        if self.mode == "single":
            SingleModeSweep.configure_sweep(self, vna, ifbw_hz)
        else:
            ContinuousModeSweep.configure_sweep(self, vna, ifbw_hz)

    def run_sweeps(self, vna, ifbw_hz):
        """Dispatch to _single_sweep_loop or _continuous_sweep_loop."""
        if self.mode == "single":
            return self._single_sweep_loop(vna, ifbw_hz)
        else:
            return self._continuous_sweep_loop(vna, ifbw_hz)

    def pre_loop_reset(self, vna):
        """Dispatch to the correct parent's pre_loop_reset."""
        if self.mode == "single":
            SingleModeSweep.pre_loop_reset(self, vna)
        else:
            ContinuousModeSweep.pre_loop_reset(self, vna)

    def post_loop_teardown(self, vna):
        """Dispatch to the correct parent's post_loop_teardown."""
        if self.mode == "single":
            SingleModeSweep.post_loop_teardown(self, vna)
        else:
            ContinuousModeSweep.post_loop_teardown(self, vna)


# ===========================================================================
# argparse entry point
# ===========================================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="LibreVNA GUI mode sweep test")
    parser.add_argument(
        "--config",
        default=os.path.join(SCRIPT_DIR, "sweep_config.yaml"),
        help="Path to YAML config (default: sweep_config.yaml in scripts/)",
    )
    parser.add_argument(
        "--mode",
        choices=["single", "continuous"],
        default="single",
        help="Sweep mode (default: single)",
    )
    parser.add_argument(
        "--no-summary",
        action="store_true",
        help="Suppress console summary",
    )
    parser.add_argument(
        "--no-save",
        action="store_true",
        help="Skip xlsx export",
    )
    args = parser.parse_args()

    test = VNAGUIModeSweepTest(
        config_path=args.config,
        mode=args.mode,
        summary=not args.no_summary,
        save_data=not args.no_save,
    )
    test.run()

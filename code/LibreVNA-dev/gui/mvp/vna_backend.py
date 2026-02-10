"""
Standalone VNA sweep backend for LibreVNA GUI.

Extracted from scripts/6_librevna_gui_mode_sweep_test.py for deployment independence.
Provides BaseVNASweep and ContinuousModeSweep classes with SCPI-based sweep control.
"""

import sys
import os
import platform
import math
import time
import json
import socket
import subprocess
import threading
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from datetime import datetime

import yaml
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .libreVNA import libreVNA

# ---------------------------------------------------------------------------
# Module-relative paths
# ---------------------------------------------------------------------------

_MODULE_DIR = os.path.dirname(os.path.abspath(__file__))

# OS-dependent GUI binary path
if platform.system() == "Windows":
    GUI_BINARY = os.path.normpath(
        os.path.join(_MODULE_DIR, "..", "..", "tools", "LibreVNA-GUI", "release", "LibreVNA-GUI.exe")
    )
else:
    GUI_BINARY = os.path.normpath(
        os.path.join(_MODULE_DIR, "..", "..", "tools", "LibreVNA-GUI")
    )

# SCPI connection
SCPI_HOST = "localhost"
SCPI_PORT = 19542
GUI_START_TIMEOUT_S = 30.0

# Polling/timing
POLL_INTERVAL_S = 0.01
SWEEP_TIMEOUT_S = 60.0

# Streaming (continuous mode)
STREAMING_PORT = 19001
CONTINUOUS_TIMEOUT_S = 300

# ---------------------------------------------------------------------------
# Console-output helpers
# ---------------------------------------------------------------------------


def _section(title):
    """Print a dashed section header."""
    width = 70
    print("\n" + "=" * width)
    print("  " + title)
    print("=" * width)


def _subsection(title):
    """Print a lighter sub-header."""
    print("\n  --- " + title + " ---")


# ---------------------------------------------------------------------------
# SweepResult dataclass
# ---------------------------------------------------------------------------


@dataclass
class SweepResult:
    """Container for one complete IFBW run (all sweeps at that setting)."""

    mode: str  # "single" or "continuous"
    ifbw_hz: int
    sweep_times: list  # list[float] wall-clock seconds per sweep
    all_s11_db: list  # list[list[float]] [sweep_idx][point_idx]
    freq_hz: list  # list[float] frequency axis
    noise_floor: float = 0.0  # filled by compute_metrics
    trace_jitter: float = 0.0  # filled by compute_metrics


# ===========================================================================
# BaseVNASweep  --  abstract base class
# ===========================================================================


class BaseVNASweep(ABC):
    """
    Shared lifecycle, GUI management, calibration loading, metric
    computation, console summary, and xlsx export.  Concrete sweep
    logic (configure + run) is left to the subclasses.
    """

    @staticmethod
    def parse_calibration_file(cal_file_path):
        """
        Parse a LibreVNA .cal file (JSON) and extract sweep parameters.

        The calibration file contains a 'measurements' array where each
        measurement has 'data.points' -- an array of {frequency, real, imag}
        objects.  The start frequency is the first point's frequency, the
        stop frequency is the last point's frequency, and num_points is
        the length of the points array.

        All measurements within a single .cal file share the same frequency
        grid, so only the first measurement is inspected.

        Parameters
        ----------
        cal_file_path : str
            Absolute or relative path to the .cal file.

        Returns
        -------
        dict
            Keys: 'start_frequency' (int, Hz),
                  'stop_frequency'  (int, Hz),
                  'num_points'      (int).

        Raises
        ------
        FileNotFoundError
            If the file does not exist on disk.
        ValueError
            If the file is not valid JSON, is missing required keys,
            or contains no measurement data.
        """
        # Resolve the calibration file path:
        # If cal_file_path is already absolute, use it as-is.
        # If it is relative (e.g. a bare filename like "SOLT_1_200M-250M_801pt.cal"),
        # resolve it against _MODULE_DIR (gui/mvp/) where .cal files are colocated.
        # This matches the path resolution logic in load_calibration() (lines 405-410).
        if os.path.isabs(cal_file_path):
            cal_abs = os.path.normpath(cal_file_path)
        else:
            cal_abs = os.path.normpath(
                os.path.join(_MODULE_DIR, cal_file_path)
            )

        if not os.path.isfile(cal_abs):
            raise FileNotFoundError(
                "Calibration file not found: {}".format(cal_abs)
            )

        try:
            with open(cal_abs, "r") as fh:
                cal_data = json.load(fh)
        except json.JSONDecodeError as exc:
            raise ValueError(
                "Calibration file is not valid JSON: {}: {}".format(cal_abs, exc)
            )

        # -- Validate required structure ----------------------------------------
        if "measurements" not in cal_data:
            raise ValueError(
                "Calibration file missing 'measurements' key: {}".format(cal_abs)
            )

        measurements = cal_data["measurements"]
        if not isinstance(measurements, list) or len(measurements) == 0:
            raise ValueError(
                "Calibration file has no measurements: {}".format(cal_abs)
            )

        first_meas = measurements[0]

        if "data" not in first_meas:
            raise ValueError(
                "First measurement missing 'data' key: {}".format(cal_abs)
            )

        if "points" not in first_meas["data"]:
            raise ValueError(
                "First measurement missing 'data.points' key: {}".format(cal_abs)
            )

        points = first_meas["data"]["points"]
        if not isinstance(points, list) or len(points) == 0:
            raise ValueError(
                "First measurement has no calibration points: {}".format(cal_abs)
            )

        # -- Extract frequency range and point count ----------------------------
        first_freq = points[0].get("frequency")
        last_freq = points[-1].get("frequency")

        if first_freq is None or last_freq is None:
            raise ValueError(
                "Calibration points missing 'frequency' field: {}".format(cal_abs)
            )

        start_frequency = int(round(first_freq))
        stop_frequency = int(round(last_freq))
        num_points = len(points)

        if start_frequency >= stop_frequency:
            raise ValueError(
                "Invalid frequency range in calibration file: start={} Hz >= "
                "stop={} Hz: {}".format(start_frequency, stop_frequency, cal_abs)
            )

        if num_points < 2:
            raise ValueError(
                "Calibration file has fewer than 2 points ({}): {}".format(
                    num_points, cal_abs
                )
            )

        return {
            "start_frequency": start_frequency,
            "stop_frequency": stop_frequency,
            "num_points": num_points,
        }

    def __init__(self, config_path, cal_file_path, mode, summary=True, save_data=True):
        """
        Parameters
        ----------
        config_path : str
            Absolute path to YAML sweep configuration file
        cal_file_path : str
            Absolute path to .cal calibration file
        mode : str
            "single" or "continuous"
        summary : bool
            Print PrettyTable summary
        save_data : bool
            Write xlsx output
        """
        self.cal_file_path = cal_file_path
        self.mode = mode
        self.summary = summary
        self.save_data = save_data

        # -- Extract frequency range and num_points from the .cal file ---------
        # The calibration file is the single source of truth for sweep
        # boundaries.  This prevents misconfiguration where the YAML config
        # specifies a different range than what the calibration covers.
        cal_params = self.parse_calibration_file(self.cal_file_path)
        self.start_freq_hz = cal_params["start_frequency"]
        self.stop_freq_hz = cal_params["stop_frequency"]
        self.num_points = cal_params["num_points"]

        # -- Load remaining parameters from YAML config ------------------------
        with open(config_path, "r") as fh:
            raw = yaml.safe_load(fh)

        cfg = raw["configurations"]
        tgt = raw["target"]

        self.stim_lvl_dbm = int(cfg["stim_lvl_dbm"])
        self.avg_count = int(cfg["avg_count"])
        self.num_sweeps = int(cfg["num_sweeps"])

        # Normalise ifbw_values: accept a single int OR a list of ints.
        raw_ifbw = tgt["ifbw_values"]
        if isinstance(raw_ifbw, list):
            self.ifbw_values = [int(v) for v in raw_ifbw]
        else:
            self.ifbw_values = [int(raw_ifbw)]

    # -----------------------------------------------------------------------
    # GUI lifecycle
    # -----------------------------------------------------------------------

    def start_gui(self):
        """
        Launch LibreVNA-GUI in headless mode and poll TCP port SCPI_PORT
        until the SCPI server accepts a connection.  Returns the Popen handle.
        """
        _section("STARTING LibreVNA-GUI")

        env = os.environ.copy()
        # Only set offscreen platform on Linux/macOS where it's available
        # Windows Qt builds only include qwindows.dll (requires desktop session)
        if platform.system() != "Windows":
            env["QT_QPA_PLATFORM"] = "offscreen"

        proc = subprocess.Popen(
            [GUI_BINARY, "--port", str(SCPI_PORT), "--no-gui"],
            env=env,
            cwd=_MODULE_DIR,  # CWD = gui/mvp/ so .cal filenames resolve correctly
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        print("  GUI PID         : {}".format(proc.pid))

        deadline = time.time() + GUI_START_TIMEOUT_S
        while True:
            try:
                s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                s.settimeout(0.5)
                s.connect((SCPI_HOST, SCPI_PORT))
                s.close()
                print("  SCPI server     : ready on {}:{}".format(SCPI_HOST, SCPI_PORT))
                return proc
            except (ConnectionRefusedError, OSError):
                s.close()
            if time.time() > deadline:
                proc.terminate()
                proc.wait()
                raise RuntimeError(
                    "LibreVNA-GUI did not open SCPI server on port {} "
                    "within {:.0f} s".format(SCPI_PORT, GUI_START_TIMEOUT_S)
                )
            time.sleep(0.25)

    def stop_gui(self, proc):
        """
        Terminate the GUI subprocess gracefully (SIGTERM) with a 5 s
        timeout, escalating to SIGKILL if needed.
        """
        _section("STOPPING LibreVNA-GUI")
        proc.terminate()
        try:
            proc.wait(timeout=5)
            print("  GUI terminated  : PID {} (clean)".format(proc.pid))
        except subprocess.TimeoutExpired:
            proc.kill()
            proc.wait()
            print("  GUI killed      : PID {} (forced)".format(proc.pid))

    # -----------------------------------------------------------------------
    # Device connection
    # -----------------------------------------------------------------------

    def connect_and_verify(self):
        """
        Open a TCP connection to LibreVNA-GUI and confirm a hardware device
        is attached.  Raises RuntimeError on any failure (the GUI was just
        started by this script so a connection problem is unexpected).

        Returns
        -------
        libreVNA
            Connected and verified wrapper instance.
        """
        _section("DEVICE CONNECTION")

        try:
            vna = libreVNA(host=SCPI_HOST, port=SCPI_PORT)
            print("  TCP connection  : OK  ({}:{})".format(SCPI_HOST, SCPI_PORT))
        except Exception as exc:
            raise RuntimeError(
                "Could not connect to LibreVNA-GUI at {}:{}: {}".format(
                    SCPI_HOST, SCPI_PORT, exc
                )
            )

        _subsection("*IDN? identification")
        try:
            idn_raw = vna.query("*IDN?")
            print("  Raw response    : {}".format(idn_raw))
            parts = [p.strip() for p in idn_raw.split(",")]
            labels = ["Manufacturer", "Model", "Serial (IDN)", "SW Version"]
            for label, val in zip(labels, parts):
                print("    {:<22s}: {}".format(label, val))
        except Exception as exc:
            print("  [WARN] *IDN? query failed: {}".format(exc))

        _subsection("DEVice:CONNect? -- device serial")
        try:
            dev_serial = vna.query(":DEV:CONN?")
            print("  Live serial     : {}".format(dev_serial))
            if dev_serial == "Not connected":
                raise RuntimeError(
                    "LibreVNA-GUI is not connected to any hardware device."
                )
        except RuntimeError:
            raise
        except Exception as exc:
            raise RuntimeError("DEVice:CONNect? query failed: {}".format(exc))

        return vna

    # -----------------------------------------------------------------------
    # Calibration
    # -----------------------------------------------------------------------

    def load_calibration(self, vna: libreVNA, cal_file_path: str) -> None:
        """
        Load calibration file into LibreVNA-GUI via SCPI.

        Parameters
        ----------
        vna : libreVNA
            Connected wrapper instance
        cal_file_path : str
            Absolute or relative path to .cal file

        Raises
        ------
        FileNotFoundError
            If calibration file not found on disk
        RuntimeError
            If SCPI load query fails
        """
        _section("CALIBRATION LOADING")

        # Resolve the calibration file for existence check:
        # If cal_file_path is just a filename (no directory separators), resolve
        # it relative to _MODULE_DIR (gui/mvp/) where the .cal file is colocated.
        # If it is already an absolute path, use it as-is.
        if os.path.isabs(cal_file_path):
            cal_check_path = os.path.normpath(cal_file_path)
        else:
            cal_check_path = os.path.normpath(
                os.path.join(_MODULE_DIR, cal_file_path)
            )

        print("  Cal file path   : {}".format(cal_check_path))

        if not os.path.isfile(cal_check_path):
            raise FileNotFoundError(
                f"Calibration file not found: {cal_check_path}"
            )

        print("  File exists     : YES")

        # -- VNA:CALibration:LOAD? <filename> ------------------------------------
        # Send just the filename to the SCPI command to avoid full Windows paths
        # with spaces that break SCPI parsing. The GUI subprocess CWD is set to
        # _MODULE_DIR (gui/mvp/) where the .cal file is colocated, so the
        # filename-only approach resolves correctly.
        cal_scpi_path = os.path.basename(cal_check_path)
        print("  SCPI cal path   : {}".format(cal_scpi_path))
        load_response = vna.query(":VNA:CAL:LOAD? " + cal_scpi_path)
        print("  LOAD? response  : {}".format(load_response))

        if load_response != "TRUE":
            raise RuntimeError(
                f"VNA:CALibration:LOAD? returned '{load_response}'. "
                f"Possible causes: GUI cannot access the path, or file is not valid."
            )

        # -- VNA:CALibration:ACTIVE? --------------------------------------------
        active_cal = vna.query(":VNA:CAL:ACTIVE?")
        print("  Active cal type : {}".format(active_cal))

    def enable_streaming_server(self, vna):
        """
        Ensure the VNA Calibrated Data streaming server is enabled on port 19001.

        If the streaming server is already enabled (port 19001 is listening),
        this method returns False immediately (fast path).

        If the streaming server is disabled, this method enables it via SCPI:
            :DEV:PREF StreamingServers.VNACalibratedData.enabled true
            :DEV:APPLYPREFERENCES

        Note that APPLYPREFERENCES saves the preference to disk and terminates
        the GUI.  The caller (run()) must stop the old GUI and start a new one.

        Parameters
        ----------
        vna : libreVNA
            The currently active connection (will become stale after this call).

        Returns
        -------
        bool
            True if the preference was set (caller must restart GUI).
            False if streaming was already enabled (no restart needed).
        """
        _section("STREAMING SERVER SETUP")

        # Test if streaming server is already listening
        print("  Testing port {} ...".format(STREAMING_PORT))
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.settimeout(1.0)
            s.connect((SCPI_HOST, STREAMING_PORT))
            s.close()
            print("  Status          : already enabled (fast path)")
            return False  # no restart needed
        except (ConnectionRefusedError, OSError):
            s.close()
            print("  Status          : not enabled, enabling now...")

        # Enable streaming server via SCPI
        # DEV:PREF set returns CME bit even on success -- use check=False
        print(
            "  Sending         : DEV:PREF StreamingServers.VNACalibratedData.enabled true"
        )
        vna.cmd(
            ":DEV:PREF StreamingServers.VNACalibratedData.enabled true", check=False
        )

        print("  Sending         : DEV:APPLYPREFERENCES")
        vna.cmd(":DEV:APPLYPREFERENCES", check=False)

        # APPLYPREFERENCES saves the preference to disk and terminates the GUI.
        # Give it a moment to save, then let the caller handle the restart.
        print("  Preference      : saved to disk")
        print("  [INFO] GUI will terminate now; caller must restart it")
        time.sleep(2)  # brief pause to ensure preference is saved

        return True  # caller must restart GUI

    # -----------------------------------------------------------------------
    # Abstract methods -- implemented by the single / continuous subclasses
    # -----------------------------------------------------------------------

    @abstractmethod
    def configure_sweep(self, vna, ifbw_hz):
        """Configure the VNA for a sweep at the given IFBW."""
        ...

    @abstractmethod
    def run_sweeps(self, vna, ifbw_hz):
        """
        Execute self.num_sweeps sweeps and return a SweepResult.

        Parameters
        ----------
        vna     : libreVNA
        ifbw_hz : int

        Returns
        -------
        SweepResult
        """
        ...

    # -----------------------------------------------------------------------
    # Metric computation (shared)
    # -----------------------------------------------------------------------

    def compute_metrics(self, result):
        """
        Fill result.noise_floor and result.trace_jitter in-place.

        Noise floor  = mean( per-sweep mean S11_dB across all freq points )
        Trace jitter = mean( per-point std across sweeps, ddof=1 )

        Parameters
        ----------
        result : SweepResult
            Must have all_s11_db populated.
        """
        s11_arr = np.array(result.all_s11_db)  # shape (num_sweeps, num_points)

        # Noise floor
        per_sweep_means = np.mean(s11_arr, axis=1)  # shape (num_sweeps,)
        result.noise_floor = float(np.mean(per_sweep_means))

        # Trace jitter
        per_point_stds = np.std(s11_arr, axis=0, ddof=1)  # shape (num_points,)
        result.trace_jitter = float(np.mean(per_point_stds))

    # -----------------------------------------------------------------------
    # Console summary (no PrettyTable dependency - simplified for GUI use)
    # -----------------------------------------------------------------------

    def print_summary(self, all_results):
        """
        Print a simple text summary with one row per IFBW result.

        Parameters
        ----------
        all_results : list[SweepResult]
        """
        _section("SWEEP TEST SUMMARY  ({} mode)".format(self.mode))

        print("{:<12} {:<12} {:<12} {:<12} {:<12}".format(
            "IFBW (kHz)", "Mean (s)", "Rate (Hz)", "Noise (dB)", "Jitter (dB)"
        ))
        print("-" * 60)

        for r in all_results:
            times_arr = np.array(r.sweep_times)
            mean_t = float(np.mean(times_arr))
            rate = 1.0 / mean_t if mean_t > 0 else float("inf")

            print("{:<12d} {:<12.4f} {:<12.2f} {:<12.2f} {:<12.4f}".format(
                r.ifbw_hz // 1000,
                mean_t,
                rate,
                r.noise_floor,
                r.trace_jitter,
            ))

    # -----------------------------------------------------------------------
    # xlsx export
    # -----------------------------------------------------------------------

    def save_xlsx(self, all_results, output_dir=None):
        """
        Write a multi-sheet xlsx workbook.

        Parameters
        ----------
        all_results : list[SweepResult]
        output_dir : str, optional
            Directory for xlsx. If None, uses ../../data/YYYYMMDD/

        Returns
        -------
        str
            Absolute path of the written xlsx file.
        """
        # -- Output path: output_dir or ../data/YYYYMMDD/ ---------------------
        if output_dir is None:
            today = datetime.now().strftime("%Y%m%d")
            output_dir = os.path.join(_MODULE_DIR, "..", "..", "data", today)

        os.makedirs(output_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "{}_sweep_test_{}.xlsx".format(self.mode, timestamp)
        full_path = os.path.join(output_dir, filename)

        # -- Style definitions ------------------------------------------------
        bold_font = Font(bold=True)
        title_font = Font(bold=True, size=13)
        section_font = Font(bold=True, size=12)
        header_fill = PatternFill(
            start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"
        )

        wb = Workbook()

        # ---- Sheet: Summary -------------------------------------------------
        ws = wb.active
        ws.title = "Summary"

        # Row 1: merged title
        ws.cell(
            row=1, column=1, value="VNA Sweep Test Summary -- {} mode".format(self.mode)
        )
        ws.cell(row=1, column=1).font = title_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

        # Row 2: blank (already empty)

        # Row 3: headers
        summary_headers = [
            "Mode",
            "IFBW (kHz)",
            "Mean Time (s)",
            "Mean Time (ms)",
            "Std Dev (s)",
            "Min Time (s)",
            "Max Time (s)",
            "Rate (Hz)",
            "Noise Floor (dB)",
            "Trace Jitter (dB)",
        ]
        for col_idx, hdr in enumerate(summary_headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=hdr)
            cell.font = bold_font
            cell.fill = header_fill

        # Row 4+: data
        for row_offset, r in enumerate(all_results):
            row = 4 + row_offset
            times_arr = np.array(r.sweep_times)
            mean_t = float(np.mean(times_arr))
            std_t = float(np.std(times_arr, ddof=1)) if len(times_arr) > 1 else 0.0
            min_t = float(np.min(times_arr))
            max_t = float(np.max(times_arr))
            rate = 1.0 / mean_t if mean_t > 0 else float("inf")

            values = [
                r.mode,
                r.ifbw_hz // 1000,
                round(mean_t, 4),
                round(mean_t * 1000, 4),
                round(std_t, 4),
                round(min_t, 4),
                round(max_t, 4),
                round(rate, 2),
                round(r.noise_floor, 2),
                round(r.trace_jitter, 2),
            ]
            for col_idx, val in enumerate(values, start=1):
                ws.cell(row=row, column=col_idx, value=val)

        # ---- Per-IFBW sheets ------------------------------------------------
        for r in all_results:
            ifbw_khz = r.ifbw_hz // 1000
            sheet_name = "IFBW_{}kHz".format(ifbw_khz)
            ws = wb.create_sheet(title=sheet_name)

            # ==== Config block (rows 1-9) ====================================
            ws.cell(row=1, column=1, value="Configuration")
            ws.cell(row=1, column=1).font = section_font

            config_rows = [
                ("Mode", self.mode),
                ("IFBW (kHz)", ifbw_khz),
                ("Start Freq (Hz)", self.start_freq_hz),
                ("Stop Freq (Hz)", self.stop_freq_hz),
                ("Points", self.num_points),
                ("STIM Level (dBm)", self.stim_lvl_dbm),
                ("Avg Count", self.avg_count),
                ("Num Sweeps", self.num_sweeps),
            ]
            for i, (label, value) in enumerate(config_rows):
                row = 2 + i  # rows 2-9
                ws.cell(row=row, column=1, value=label).font = bold_font
                ws.cell(row=row, column=2, value=value)

            # ==== Timing block (starts row 10) ===============================
            timing_header_row = 10
            ws.cell(row=timing_header_row, column=1, value="Timing")
            ws.cell(row=timing_header_row, column=1).font = section_font

            timing_cols = [
                "Sweep #",
                "Sweep Time (s)",
                "Sweep Time (ms)",
                "Update Rate (Hz)",
            ]
            col_header_row = timing_header_row + 1  # row 11
            for col_idx, hdr in enumerate(timing_cols, start=1):
                cell = ws.cell(row=col_header_row, column=col_idx, value=hdr)
                cell.font = bold_font
                cell.fill = header_fill

            for sweep_idx, t in enumerate(r.sweep_times):
                row = col_header_row + 1 + sweep_idx  # row 12+
                rate = 1.0 / t if t > 0 else float("inf")
                ws.cell(row=row, column=1, value=sweep_idx + 1)
                ws.cell(row=row, column=2, value=round(t, 4))
                ws.cell(row=row, column=3, value=round(t * 1000, 4))
                ws.cell(row=row, column=4, value=round(rate, 2))

            # blank row after timing
            traces_section_row = col_header_row + 1 + len(r.sweep_times) + 1

            # ==== S11 Traces block ============================================
            ws.cell(row=traces_section_row, column=1, value="S11 Traces")
            ws.cell(row=traces_section_row, column=1).font = section_font

            trace_col_header_row = traces_section_row + 1
            # Column headers: Frequency (Hz) | Sweep_1 S11 (dB) | ...
            ws.cell(row=trace_col_header_row, column=1, value="Frequency (Hz)")
            ws.cell(row=trace_col_header_row, column=1).font = bold_font
            ws.cell(row=trace_col_header_row, column=1).fill = header_fill
            for s_idx in range(len(r.all_s11_db)):
                col = 2 + s_idx
                ws.cell(
                    row=trace_col_header_row,
                    column=col,
                    value="Sweep_{} S11 (dB)".format(s_idx + 1),
                )
                ws.cell(row=trace_col_header_row, column=col).font = bold_font
                ws.cell(row=trace_col_header_row, column=col).fill = header_fill

            # Data rows: one per frequency point
            for pt_idx, freq in enumerate(r.freq_hz):
                row = trace_col_header_row + 1 + pt_idx
                ws.cell(row=row, column=1, value=freq)
                for s_idx in range(len(r.all_s11_db)):
                    ws.cell(
                        row=row,
                        column=2 + s_idx,
                        value=round(r.all_s11_db[s_idx][pt_idx], 4),
                    )

            # blank row after traces
            metrics_section_row = trace_col_header_row + 1 + len(r.freq_hz) + 1

            # ==== Metrics block ===============================================
            ws.cell(row=metrics_section_row, column=1, value="Metrics")
            ws.cell(row=metrics_section_row, column=1).font = section_font

            ws.cell(row=metrics_section_row + 1, column=1, value="Noise Floor (dB)")
            ws.cell(row=metrics_section_row + 1, column=1).font = bold_font
            ws.cell(
                row=metrics_section_row + 1, column=2, value=round(r.noise_floor, 4)
            )

            ws.cell(row=metrics_section_row + 2, column=1, value="Trace Jitter (dB)")
            ws.cell(row=metrics_section_row + 2, column=1).font = bold_font
            ws.cell(
                row=metrics_section_row + 2, column=2, value=round(r.trace_jitter, 6)
            )

        # -- Write workbook ---------------------------------------------------
        wb.save(full_path)
        return full_path

    # -----------------------------------------------------------------------
    # Lifecycle hooks  --  override in subclasses for one-time setup / teardown
    # -----------------------------------------------------------------------

    def pre_loop_reset(self, vna):
        """Called once before the IFBW loop.  Override for mode-specific one-time setup."""
        pass

    def post_loop_teardown(self, vna):
        """Called once after the IFBW loop completes.  Override for cleanup."""
        pass

    # -----------------------------------------------------------------------
    # Main orchestration
    # -----------------------------------------------------------------------

    def run(self):
        """
        Top-level entry point.

        Sequence
        --------
        1.  start_gui()
        2.  connect_and_verify()
        3.  enable_streaming_server() (continuous mode only; auto-enables + may restart GUI)
        4.  load_calibration()
        5.  pre_loop_reset()   (one-time mode-specific setup)
        6.  For each IFBW: configure_sweep() -> run_sweeps() -> compute_metrics()
        7.  post_loop_teardown()   (one-time mode-specific cleanup)
        8.  print_summary()  (if self.summary)
        9.  save_xlsx()      (if self.save_data)
        10. stop_gui()       (in finally block)
        """
        gui_proc = self.start_gui()
        try:
            vna = self.connect_and_verify()

            # If continuous mode, enable streaming server (may restart GUI)
            if self.mode == "continuous":
                needs_restart = self.enable_streaming_server(vna)
                if needs_restart:
                    # APPLYPREFERENCES terminated the old GUI; restart it
                    _section("RESTARTING GUI")
                    self.stop_gui(gui_proc)
                    gui_proc = self.start_gui()
                    vna = self.connect_and_verify()

            self.load_calibration(vna, self.cal_file_path)

            all_results = []

            self.pre_loop_reset(vna)

            for ifbw_hz in self.ifbw_values:
                _section(
                    "IFBW = {} kHz  --  {} mode".format(ifbw_hz // 1000, self.mode)
                )

                self.configure_sweep(vna, ifbw_hz)
                result = self.run_sweeps(vna, ifbw_hz)
                self.compute_metrics(result)

                # Per-IFBW progress
                _subsection("IFBW {} kHz -- metrics".format(ifbw_hz // 1000))
                times_arr = np.array(result.sweep_times)
                print(
                    "    Mean sweep time : {:.4f} s".format(float(np.mean(times_arr)))
                )
                print(
                    "    Update rate     : {:.2f} Hz".format(
                        1.0 / float(np.mean(times_arr))
                        if float(np.mean(times_arr)) > 0
                        else float("inf")
                    )
                )
                print("    Noise floor     : {:.2f} dB".format(result.noise_floor))
                print("    Trace jitter    : {:.4f} dB".format(result.trace_jitter))

                all_results.append(result)

            self.post_loop_teardown(vna)

            # -- Console summary ----------------------------------------------
            if self.summary:
                self.print_summary(all_results)

            # -- xlsx export --------------------------------------------------
            if self.save_data:
                _section("SAVING RESULTS")
                xlsx_path = self.save_xlsx(all_results)
                print("  xlsx written    : {}".format(xlsx_path))

            print()  # trailing blank line

        finally:
            self.stop_gui(gui_proc)


# ===========================================================================
# ContinuousModeSweep
# ===========================================================================


class ContinuousModeSweep(BaseVNASweep):
    """
    Continuous-sweep mode: streaming callback on port 19001.

    The streaming callback is registered ONCE in pre_loop_reset() and
    removed ONCE in post_loop_teardown().  Between IFBWs the callback
    stays connected; only the mutable state it reads is swapped.  This
    avoids the remove + re-add path that triggers a known bug in
    libreVNA.py line 148 (the thread-exit guard checks the wrong list
    length).

    configure_sweep sends the full configuration INCLUDING STOP (no
    ACQ:STOP / SINGLE TRUE prefix -- those are handled by
    pre_loop_reset / post_loop_teardown).
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._state_holder = [None]  # mutable container; callback always reads [0]
        self._stream_callback = None  # the single registered callback closure

    # -- Inner helper: shared mutable state guarded by a Lock -----------------

    class _SweepState:
        """
        All mutable state touched by the streaming callback, bundled into a
        single object so that the lock scope is unambiguous.
        """

        def __init__(self, num_points, num_sweeps):
            self.lock = threading.Lock()
            self.done_event = threading.Event()
            self.num_points = num_points
            self.num_sweeps = num_sweeps
            self.sweep_count = 0
            self.sweep_start_time = 0.0
            self.sweep_end_times = []  # list[float]
            self.sweep_start_times = []  # list[float]
            self.current_s11 = []  # list[complex], current sweep
            self.all_s11_complex = []  # list[list[complex]], all completed sweeps

    # -----------------------------------------------------------------------

    def configure_sweep(self, vna, ifbw_hz):
        """
        Continuous-mode configuration -- full sequence including STOP.

        SCPI sequence
        -------------
        :DEV:MODE VNA
        :VNA:SWEEP FREQUENCY
        :VNA:STIM:LVL <dBm>
        :VNA:ACQ:IFBW  <Hz>
        :VNA:ACQ:AVG   <n>
        :VNA:ACQ:POINTS <n>
        :VNA:FREQuency:START <Hz>
        :VNA:FREQuency:STOP  <Hz>   -- included because STOP is NOT the trigger here
        """
        _subsection(
            "Continuous-mode configuration  (IFBW = {} kHz)".format(ifbw_hz // 1000)
        )

        vna.cmd(":DEV:MODE VNA")
        print("  Mode            : VNA")

        vna.cmd(":VNA:SWEEP FREQUENCY")
        print("  Sweep type      : FREQUENCY")

        vna.cmd(":VNA:STIM:LVL {}".format(self.stim_lvl_dbm))
        print("  Stimulus level  : {} dBm".format(self.stim_lvl_dbm))

        vna.cmd(":VNA:ACQ:IFBW {}".format(ifbw_hz))
        print("  IF bandwidth    : {} Hz  ({} kHz)".format(ifbw_hz, ifbw_hz // 1000))

        vna.cmd(":VNA:ACQ:AVG {}".format(self.avg_count))
        print("  Averaging       : {} sweep(s)".format(self.avg_count))

        vna.cmd(":VNA:ACQ:POINTS {}".format(self.num_points))
        print("  Points          : {}".format(self.num_points))

        vna.cmd(":VNA:FREQuency:START {}".format(self.start_freq_hz))
        print(
            "  Start freq      : {} Hz  ({:.3f} GHz)".format(
                self.start_freq_hz, self.start_freq_hz / 1e9
            )
        )

        vna.cmd(":VNA:FREQuency:STOP {}".format(self.stop_freq_hz))
        print(
            "  Stop freq       : {} Hz  ({:.3f} GHz)".format(
                self.stop_freq_hz, self.stop_freq_hz / 1e9
            )
        )

    def run_sweeps(self, vna, ifbw_hz):
        """Dispatch to the streaming callback loop."""
        return self._continuous_sweep_loop(vna, ifbw_hz)

    def _make_callback(self, state_holder):
        """
        Return a closure that reads its state from state_holder[0].

        state_holder is a one-element list.  The main thread swaps [0] to a
        fresh _SweepState between IFBWs; the callback picks up the new state
        on the next point without any remove/re-add of the TCP connection.
        """

        def _callback(data):
            state = state_holder[0]
            if state is None:
                return  # gap during state swap -- drop silently
            if "Z0" not in data:
                return

            point_num = data["pointNum"]
            s11_complex = data["measurements"].get("S11", complex(0, 0))

            with state.lock:
                if point_num == 0:
                    state.sweep_start_time = time.time()
                    state.current_s11 = []

                state.current_s11.append(s11_complex)

                if point_num == state.num_points - 1:
                    sweep_end = time.time()
                    state.sweep_end_times.append(sweep_end)
                    state.sweep_start_times.append(state.sweep_start_time)
                    state.all_s11_complex.append(list(state.current_s11))
                    state.sweep_count += 1

                    if state.sweep_count >= state.num_sweeps:
                        state.done_event.set()

        return _callback

    def pre_loop_reset(self, vna):
        """
        Register streaming callback ONCE and prepare continuous mode.

        The streaming server is guaranteed to be enabled by run() before
        this method is called, so we can directly connect without error
        handling.
        """
        _subsection("Continuous-mode streaming setup (one-time)")

        # Create the persistent callback referencing the mutable state holder
        self._stream_callback = self._make_callback(self._state_holder)

        # Stop any residual acquisition before connecting
        vna.cmd(":VNA:ACQ:STOP")
        print("  Pre-stop        : sent")

        # Register the streaming callback (streaming server is guaranteed enabled)
        vna.add_live_callback(STREAMING_PORT, self._stream_callback)
        print(
            "  Streaming       : callback registered on port {}".format(STREAMING_PORT)
        )

    def post_loop_teardown(self, vna):
        """Stop acquisition, restore single mode, remove the streaming callback."""
        _subsection("Continuous-mode streaming teardown")

        vna.cmd(":VNA:ACQ:STOP")
        print("  Acquisition     : stopped")

        vna.cmd(":VNA:ACQ:SINGLE TRUE")
        print("  Sweep mode      : restored to SINGLE (TRUE)")

        if self._stream_callback is not None:
            vna.remove_live_callback(STREAMING_PORT, self._stream_callback)
            self._stream_callback = None
            print("  Streaming       : callback removed")

    def _continuous_sweep_loop(self, vna, ifbw_hz):
        """
        Run self.num_sweeps continuous sweeps for one IFBW value.

        The streaming callback is already registered (by pre_loop_reset).
        This method only manages per-IFBW acquisition start/stop and state.

        Sequence per IFBW
        ------------------
        1.  ACQ:STOP              -- halt any sweeps from the previous IFBW
        2.  sleep 0.1 s           -- drain buffered streaming points so the
                                     fresh state does not see stale data
        3.  swap state_holder[0]  -- atomically switch the callback to a
                                     new _SweepState (GIL makes list-item
                                     assignment atomic)
        4.  ACQ:SINGLE FALSE      -- ensure continuous mode (idempotent)
        5.  ACQ:RUN               -- start back-to-back sweeps
        6.  wait on done_event    -- callback sets it when sweep_count reaches target
        7.  ACQ:STOP              -- halt after target is reached

        Returns
        -------
        SweepResult
        """
        _subsection(
            "Continuous-sweep loop  ({} sweeps, IFBW {} kHz)".format(
                self.num_sweeps, ifbw_hz // 1000
            )
        )

        # 1. Stop previous IFBW's sweeps
        vna.cmd(":VNA:ACQ:STOP")
        print("  ACQ:STOP        : sent (halt previous IFBW)")

        # 2. Drain any buffered streaming points from the previous sweep
        time.sleep(0.1)

        # 3. Swap in a fresh state -- the callback picks it up on next point
        fresh_state = self._SweepState(self.num_points, self.num_sweeps)
        self._state_holder[0] = fresh_state
        print("  State           : fresh _SweepState installed")

        # 4. Ensure continuous mode (ACQ:STOP does NOT change SINGLE setting,
        #    but belt-and-suspenders is cheap here)
        vna.cmd(":VNA:ACQ:SINGLE FALSE")
        print("  Sweep mode      : CONTINUOUS  (SINGLE FALSE)")

        # 5. Start acquisition
        vna.cmd(":VNA:ACQ:RUN")
        print("  Acquisition     : started (continuous)")
        print(
            "  Collecting {} sweeps via streaming callback ...".format(self.num_sweeps)
        )

        # 6. Wait for completion
        completed = fresh_state.done_event.wait(timeout=CONTINUOUS_TIMEOUT_S)

        if not completed:
            # Tear down before raising
            vna.cmd(":VNA:ACQ:STOP")
            raise TimeoutError(
                "Only {}/{} sweeps received within {} s timeout.  "
                "Check streaming server connectivity.".format(
                    fresh_state.sweep_count, self.num_sweeps, CONTINUOUS_TIMEOUT_S
                )
            )

        # 7. Stop sweeps for this IFBW (do NOT restore SINGLE TRUE here --
        #    that is post_loop_teardown's job)
        vna.cmd(":VNA:ACQ:STOP")
        print("  Acquisition     : stopped")

        # -- Progress printout ---------------------------------------------------
        with fresh_state.lock:
            for i in range(fresh_state.sweep_count):
                duration = (
                    fresh_state.sweep_end_times[i] - fresh_state.sweep_start_times[i]
                )
                dur_rate = 1.0 / duration if duration > 0 else float("inf")
                if i == 0:
                    print(
                        "    Sweep {:>2d}/{:<2d}  :  dur {:.4f} s  ({:.1f} Hz)  "
                        "inter-sweep: --".format(
                            i + 1, fresh_state.sweep_count, duration, dur_rate
                        )
                    )
                else:
                    interval = (
                        fresh_state.sweep_end_times[i]
                        - fresh_state.sweep_end_times[i - 1]
                    )
                    int_rate = 1.0 / interval if interval > 0 else float("inf")
                    print(
                        "    Sweep {:>2d}/{:<2d}  :  dur {:.4f} s  ({:.1f} Hz)  "
                        "inter-sweep: {:.4f} s  ({:.1f} Hz)".format(
                            i + 1,
                            fresh_state.sweep_count,
                            duration,
                            dur_rate,
                            interval,
                            int_rate,
                        )
                    )

        # -- Convert all sweeps to dB -------------------------------------------
        freq_hz = list(
            np.linspace(
                float(self.start_freq_hz), float(self.stop_freq_hz), self.num_points
            )
        )

        all_s11_db = []
        sweep_times = []

        for i in range(fresh_state.sweep_count):
            t = fresh_state.sweep_end_times[i] - fresh_state.sweep_start_times[i]
            sweep_times.append(t)

            raw_s11 = fresh_state.all_s11_complex[i]
            s11_db_list = []
            for gamma in raw_s11:
                magnitude = abs(gamma)
                if magnitude < 1e-12:
                    magnitude = 1e-12
                s11_db_list.append(20.0 * math.log10(magnitude))
            all_s11_db.append(s11_db_list)

        return SweepResult(
            mode="continuous",
            ifbw_hz=ifbw_hz,
            sweep_times=sweep_times,
            all_s11_db=all_s11_db,
            freq_hz=freq_hz,
        )

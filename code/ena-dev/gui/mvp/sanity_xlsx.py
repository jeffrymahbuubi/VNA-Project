"""sanity_xlsx.py — Device Sanity Check xlsx writer.

Writes the per-IFBW benchmark summary (mean sweep time, update rate, noise floor,
trace jitter) the GUI's Sanity Check mode collects, in a multi-block sheet
echoing the LibreVNA bench schema (Configuration + Metrics). Secondary mode —
the Monitor CSV (dataflux.py) is the primary objective output.
"""

from __future__ import annotations

import os
from typing import Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font


def write_sanity_xlsx(
    rows: Sequence[dict],          # per-IFBW: ifbw_khz, mean_ms, rate_hz, nf_db, jitter_db
    *,
    meta: dict,                    # model/serial/start_mhz/stop_mhz/points/power_dbm/num_sweeps
    out_dir: str,
    filename: str,
) -> Optional[str]:
    if not rows:
        return None
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sanity Check"
    bold = Font(bold=True)

    # Configuration block
    ws["A1"] = "Configuration"; ws["A1"].font = bold
    cfg = [
        ("Model", meta.get("model", "E5063A")),
        ("Serial", meta.get("serial", "")),
        ("Start (MHz)", meta.get("start_mhz")),
        ("Stop (MHz)", meta.get("stop_mhz")),
        ("Points", meta.get("points")),
        ("Power (dBm)", meta.get("power_dbm")),
        ("Sweeps / IFBW", meta.get("num_sweeps")),
        ("Parameter", "S11"),
    ]
    r = 2
    for k, v in cfg:
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
        r += 1

    # Metrics table
    r += 1
    ws.cell(row=r, column=1, value="Metrics (per IFBW)").font = bold
    r += 1
    headers = ["IFBW (kHz)", "Mean Sweep (ms)", "Update Rate (Hz)",
               "Noise Floor (dB)", "Trace Jitter (dB)"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h).font = bold
    r += 1
    for row in rows:
        ws.cell(row=r, column=1, value=round(row["ifbw_khz"], 3))
        ws.cell(row=r, column=2, value=round(row["mean_ms"], 3))
        ws.cell(row=r, column=3, value=round(row["rate_hz"], 3))
        ws.cell(row=r, column=4, value=round(row["nf_db"], 3))
        ws.cell(row=r, column=5, value=round(row["jitter_db"], 4))
        r += 1

    for col, width in zip("ABCDE", (16, 16, 16, 16, 16)):
        ws.column_dimensions[col].width = width

    wb.save(path)
    return path
